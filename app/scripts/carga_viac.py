import mysql.connector
import os
import sys
import openpyxl
from openpyxl.styles import  PatternFill , Border, Font, Alignment, Side
from openpyxl.utils import get_column_letter
from urllib.parse import urlparse
import pandas as pd
import math
from openpyxl import Workbook
from openpyxl.drawing.image import Image

db_url = 'mysql://root:jhGAlwlZnNpfenELaVxNKrEcWOcfMuFY@caboose.proxy.rlwy.net:14619/railway'
parsed_url = urlparse(db_url)
db_config = {
    'host': parsed_url.hostname,
    'user': parsed_url.username,
    'password': parsed_url.password,
    'database': parsed_url.path[1:],
    'port': parsed_url.port
}

conn = mysql.connector.connect(**db_config)
cursor = conn.cursor(dictionary=True)

# ===============================
#  CONSULTAS
# ===============================
cargas_query = """
SELECT 
    docentes.nombres,
    docentes.apellidos,
    cargo.nombre AS cargo,
    categoria.nombre AS categoria,
    cursos.nombre AS curso,
    cursos.ciclo,
    cursos.n_grupos,
    cursos.n_estudiantes,
    cargas.total_horas,
    cargas.horas_t_carga,
    t_practica.nombre AS tipo_practica,
    cargas.horas_p_carga

FROM 
    cargas
INNER JOIN docentes ON docentes.id = cargas.docente_id
INNER JOIN cargo ON cargo.id = docentes.cargo_id
INNER JOIN categoria ON categoria.id = docentes.categoria_id
INNER JOIN cursos ON cursos.id = cargas.curso_id
INNER JOIN t_practica ON t_practica.id = cursos.t_practica_id;

"""

from openpyxl.styles import PatternFill

# Ejemplo de colores suaves (puedes ajustar el código del color si deseas otro tono)
amarillo_suave = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Amarillo pastel
celeste_suave = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")  # Celeste pastel
verde_suave = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")    # Verde pastel

# Supongamos que estás escribiendo el nombre del docente en la fila row_num
row_num = 10  # Ejemplo
fill_color = amarillo_suave



cursor.execute(cargas_query)
cargas_data = cursor.fetchall()

docentes = {}
for row in cargas_data:
    key = (row['nombres'], row['apellidos'], row['cargo'], row['categoria'])
    docentes.setdefault(key, []).append(row)

# Crear nuevo libro y hoja
def crear_hoja(wb, nombre_hoja, docentes_filtrados):
    ws = wb.create_sheet(nombre_hoja)

    for item in cargas_data:
        item['ciclo'] = str(item['ciclo'])
        item['n_grupos'] = str(item['n_grupos'])
        item['n_estudiantes'] = str(item['n_estudiantes'])
        item['total_horas'] = str(item['total_horas'])


    df = pd.DataFrame(cargas_data)

    # ======== CONFIGURACIÓN DE ANCHOS ========
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 60
    ws.row_dimensions[2].height = 50

    # Columnas C hasta U
    for col in range(ord('C'), ord('U') + 1):
        ws.column_dimensions[chr(col)].width = 6

    # ======== CONFIGURACIÓN DE FILAS ========
    ws.row_dimensions[11].height = 75

    for col in range(1, 12 + 1):  # A = 1, L = 12
        col_letter = get_column_letter(col)
        ws.merge_cells(f"{col_letter}9:{col_letter}11")

    # Ejemplo: texto centrado en el rango combinado
    ws['A9'] = "Nº"
    ws['A9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['A9'].font = Font(bold=True, size=10)
    
    ws['B9'] = "DOCENTE/ASIGNATURA"
    ws['B9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['B9'].font = Font(bold=True, size=10)

    ws['C9'] = "CARGO"
    ws['C9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    ws['C9'].font = Font(bold=True, size=10)

    ws['D9'] = "CATEGORIA/ DEDICACION"
    ws['D9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    ws['D9'].font = Font(bold=True, size=10)

    ws['E9'] = "ESCUELA EN LA QUE SE DICTA"
    ws['E9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    ws['E9'].font = Font(bold=True, size=10)

    ws['F9'] = "CICLO"
    ws['F9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    ws['F9'].font = Font(bold=True, size=10)

    ws['G9'] = "PERIODO SEMESTRAL"
    ws['G9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    ws['G9'].font = Font(bold=True, size=10)

    ws['H9'] = "TURNO"
    ws['H9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    ws['H9'].font = Font(bold=True, size=10)

    ws['I9'] = "SECCION"
    ws['I9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    ws['I9'].font = Font(bold=True, size=10)

    ws['J9'] = "RESPONSABLE DEL CURSO"
    ws['J9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    ws['J9'].font = Font(bold=True, size=10)

    ws['K9'] = "NÚMERO DE ESTUDIANTES"
    ws['K9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    ws['K9'].font = Font(bold=True, size=10)

    ws['L9'] = "ESTUDIANTES POR GRUPO DE PRACTICAS"
    ws['L9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    ws['L9'].font = Font(bold=True, size=10)

    ws['M9'] = "DISTRIBUCION DE HORAS"
    ws['M9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['M9'].font = Font(bold=True, size=10)

    ws['M10'] = "TOTAL"
    ws['M10'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    ws['M10'].font = Font(bold=True, size=10)

    ws['N10'] = "TEORÍA"
    ws['N10'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    ws['N10'].font = Font(bold=True, size=10)

    ws['O10'] = "PRÁCTICA"
    ws['O10'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['O10'].font = Font(bold=True, size=10)

    ws['O11'] = "CLÍNICA"
    ws['O11'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    ws['O11'].font = Font(bold=True, size=10)

    ws['P11'] = "LABORATORIO"
    ws['P11'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    ws['P11'].font = Font(bold=True, size=10)

    ws['Q11'] = "GABINETE"
    ws['Q11'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    ws['Q11'].font = Font(bold=True, size=10)

    ws['R11'] = "CAMPO"
    ws['R11'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    ws['R11'].font = Font(bold=True, size=10)

    ws['S11'] = "AULA"
    ws['S11'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    ws['S11'].font = Font(bold=True, size=10)

    ws['T11'] = "# SEMESTRE"
    ws['T11'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    ws['T11'].font = Font(bold=True, size=10)

    ws['U11'] = "TOTAL CARGA "
    ws['U11'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
    ws['U11'].font = Font(bold=True, size=10)

    ws['T9'] = "TOTAL DE HORAS"
    ws['T9'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['T9'].font = Font(bold=True, size=10)

    ws['B5'] = "FACULTAD"
    ws['B5'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws['B5'].font = Font(bold=True, size=10)

    ws['B6'] = "DEPARTAMENTO/ÁREA ACADÉMICA"
    ws['B6'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws['B6'].font = Font(bold=True, size=10)

    ws['D5'] = "CIENCIAS DE LA SALUD"
    ws['D5'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws['D5'].font = Font(bold=True, size=10)

    ws['D6'] = "DEPARTAMENTO ACADÉMICO DE ENFERMERÍA"
    ws['D6'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws['D6'].font = Font(bold=True, size=10)



    ws['A2'] = "DISTRIBUCIÓN  DE LA CARGA LECTIVA - AÑO ACADÉMICO 20## - # SEMESTRE"
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws['A2'].font = Font(bold=True, size=10)


    ws.merge_cells('A2:U2')
    ws.merge_cells('M10:M11')
    ws.merge_cells('N10:N11')
    ws.merge_cells('T9:U10')
    ws.merge_cells('O10:S10')
    ws.merge_cells('M9:S9')

    ws.merge_cells('A8:U8')
    ws.merge_cells('D5:U5')
    ws.merge_cells('D6:U6')
    ws.merge_cells('D7:U7')

    def aplicar_color(ws, rango, fill):
        for row in ws[rango]:
            for cell in row:
                cell.fill = fill

    # Aplicar colores por rango
    aplicar_color(ws, 'A9:D11', amarillo_suave)
    aplicar_color(ws, 'E9:J11', celeste_suave)
    aplicar_color(ws, 'K9:L11', amarillo_suave)
    aplicar_color(ws, 'M9:S11', verde_suave)
    aplicar_color(ws, 'T9:U11', amarillo_suave)

    #/*img_path = os.path.join(os.getcwd(), "1esen.jpg")
    #img_path = os.path.abspath(img_path)

    #img = Image(img_path)
    #img.width = 70
    #img.height = 70
    #ws.add_image(img, "B2")

    #img_path = os.path.join(os.getcwd(), "2unjbg.jpg")
    #img_path = os.path.abspath(img_path)

    #img2 = Image(img_path)
    #img2.width = 70
    #img2.height = 70
    #ws.add_image(img2, "T2")


    row_num = 12
    count = 1

    for docente, cursos in docentes_filtrados.items():
        nombres, apellidos, cargo, categoria = docente
        total_carga = sum(int(c['total_horas']) for c in cursos if c['total_horas'] is not None and str(c['total_horas']).isdigit())


        # Fila principal del docente
        ws[f"A{row_num}"] = count
        ws[f"B{row_num}"] = f"{nombres} {apellidos}"
        ws[f"C{row_num}"] = cargo
        ws[f"D{row_num}"] = categoria
        ws[f"E{row_num}"] = "ESEN"
        ws[f"M{row_num}"] = total_carga
        ws[f"U{row_num}"] = total_carga
        
        # ---- APLICAR COLOR A TODA LA FILA DEL DOCENTE (A hasta U) ----
        for col in range(ord('A'), ord('U') + 1):
            ws[f"{chr(col)}{row_num}"].fill = celeste_suave

        # Centrar columnas C–U en la fila del docente
        for col in range(ord('C'), ord('U') + 1):
            cell = ws[f"{chr(col)}{row_num}"]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Cursos del docente
        for c in cursos:
            row_num += 1
            curso = c['curso']
            ciclo = int(c['ciclo'])
            n_grupos = int(c['n_grupos'])
            n_est = int(c['n_estudiantes'])
            total_horas = c['total_horas']
            horas_t = c['horas_t_carga']
            horas_p = c['horas_p_carga']
            tipo = c['tipo_practica'].lower()

            ws[f"B{row_num}"] = curso
            ws[f"F{row_num}"] = ciclo
            ws[f"G{row_num}"] = "II-S" if ciclo % 2 == 0 else "I-S"
            ws[f"H{row_num}"] = "M/T"
            ws[f"I{row_num}"] = "A" if "- A" in curso else ("B" if "- B" in curso else "")
            ws[f"K{row_num}"] = n_est
            ws[f"L{row_num}"] = math.ceil(n_est / n_grupos)
            ws[f"M{row_num}"] = total_horas
            ws[f"N{row_num}"] = horas_t

            # Ubicar horas de práctica según tipo
            practicas_cols = {'clinica': 'O', 'laboratorio': 'P', 'gabinete': 'Q', 'campo': 'R', 'aula': 'S'}
            if tipo in practicas_cols:
                ws[f"{practicas_cols[tipo]}{row_num}"] = horas_p

            
            # === Centrar todas las celdas de C a U ===
            # Centrar todas las columnas de C a U del docente principal
            for col in range(ord('C'), ord('U') + 1):
                ws[f"{chr(col)}{row_num}"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


        row_num += 1
        count += 1

    ws.column_dimensions['C'].width = 20
    
    borde_fino = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    # Aplicar bordes desde A12 hasta U<última fila>
    ultima_fila = ws.max_row  # Detecta la última fila con datos

    for row in ws.iter_rows(min_row=9, max_row=ultima_fila, min_col=1, max_col=21):  # Columna U = 21
        for cell in row:
            cell.border = borde_fino
    
    return ws


wb = Workbook()
wb.remove(wb.active)  # Eliminar hoja vacía por defecto

# Filtrar docentes
docentes_docentes = {k: v for k, v in docentes.items() if k[2] != 'JP'}
docentes_jp = {k: v for k, v in docentes.items() if k[2] == 'JP'}

# Crear hojas
crear_hoja(wb, "DOCENTES", docentes_docentes)
crear_hoja(wb, "JEFE DE PRACTICA CONTRATO", docentes_jp)

# Guardar
wb.save("CARGA_VIAC.xlsx")
print("Archivo 'CARGA_VIAC.xlsx' creado con ambas hojas correctamente.")