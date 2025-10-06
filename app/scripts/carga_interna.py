import mysql.connector
import openpyxl
from openpyxl.styles import Border,PatternFill, Font, Alignment, Side
from urllib.parse import urlparse
import pandas as pd
import math
import os


# ===============================
#  CONFIGURACIÓN Y CONEXIÓN
# ===============================
colores = [
    "FFDDC1", "FFABAB", "FFC3A0", "FF8C94",
    "FF6F59", "FF5C39", "FF3E00", "FF3300",
    "FF1F00", "FF0000", "D80000", "B70000"
]

db_url = 'mysql://root:jhGAlwlZnNpfenELaVxNKrEcWOcfMuFY@caboose.proxy.rlwy.net:14619/railway'
parsed_url = urlparse(db_url)
db_config = {
    'host': parsed_url.hostname,
    'user': parsed_url.username,
    'password': parsed_url.password,
    'database': parsed_url.path[1:],
    'port': parsed_url.port
}

conn = mysql.connector.connect(**db_config)
cursor = conn.cursor(dictionary=True)

# ===============================
#  CONSULTAS
# ===============================
cursos_query = """
SELECT DISTINCT c.id, c.nombre, c.ciclo, c.horas_t, c.horas_p,
       c.n_estudiantes, c.n_grupos
FROM cargas cg
INNER JOIN cursos c ON cg.curso_id = c.id
ORDER BY c.ciclo;
"""

docentes_query = """
SELECT DISTINCT d.id, d.nombres, d.apellidos, d.categoria_id
FROM docentes d
INNER JOIN cargas cg ON cg.docente_id = d.id;
"""

cargas_query = """
SELECT d.id AS docente_id, d.nombres, d.apellidos,
       cg.curso_id, cg.horas_t_carga, cg.horas_p_carga
FROM cargas cg
INNER JOIN docentes d ON d.id = cg.docente_id;
"""

# ===============================
#  OBTENER DATOS
# ===============================
cursor.execute(cursos_query)
cursos_data = cursor.fetchall()

cursor.execute(docentes_query)
docentes_data = cursor.fetchall()

cursor.execute(cargas_query)
cargas_data = cursor.fetchall()

# Convertir a strings para consistencia
for item in cursos_data:
    item['id'] = str(item['id'])
    item['ciclo'] = str(item['ciclo'])

for item in cargas_data:
    item['curso_id'] = str(item['curso_id'])
    item['horas_t_carga'] = str(item['horas_t_carga'])
    item['horas_p_carga'] = str(item['horas_p_carga'])

df = pd.DataFrame(cursos_data)

# ===============================
#  CREAR ARCHIVO EXCEL
# ===============================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cursos"

bold_black_font = Font(bold=True, color="000000")
ws.column_dimensions['A'].width = 5.6
ws.column_dimensions['B'].width = 62

# Encabezado principal
ws.merge_cells('A1:B2')
ws["A1"] = "DISTRIBUCIÓN DE CARGA ACADÉMICA DE DOCENTES ORDINARIOS, CONTRATADOS Y JEFES DE PRÁCTICA TP/20 HORAS CONTRATADOS - ESCUELA PROFESIONAL DE ENFERMERÍA II SEMESTRE 2025"
ws["A1"].font = Font(size=12, bold=True)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws["A1"].fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")

# Encabezados de ciclos y cursos
for idx, row in df.iterrows():
    col = 3 + idx
    ciclo_cell = ws.cell(row=1, column=col, value=row['ciclo'])
    ciclo_cell.font = bold_black_font
    ciclo_cell.alignment = Alignment(horizontal="center", vertical="center")
    nombre_cell = ws.cell(row=2, column=col, value=row['nombre'])
    nombre_cell.font = bold_black_font
    nombre_cell.alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
    color = colores[int(row['ciclo']) - 1]
    ciclo_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    nombre_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# Etiquetas generales
labels = ["Horas", "Estudiantes", "Grupos", "Estudiantes por grupo"]
for i, label in enumerate(labels, start=3):
    ws.merge_cells(f"A{i}:B{i}")
    ws[f"A{i}"] = label
    ws[f"A{i}"].font = bold_black_font
    ws[f"A{i}"].alignment = Alignment(horizontal="center", vertical="center")

# Datos de cursos
for idx, row in df.iterrows():
    col = 3 + idx
    ws.cell(row=3, column=col, value=f"T{row['horas_t']} P{row['horas_p']}").alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=4, column=col, value=row['n_estudiantes']).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=5, column=col, value=row['n_grupos']).alignment = Alignment(horizontal="center", vertical="center")
    est_x_grupo = math.ceil(int(row['n_estudiantes']) / int(row['n_grupos'])) if int(row['n_grupos']) > 0 else 0
    ws.cell(row=6, column=col, value=est_x_grupo).alignment = Alignment(horizontal="center", vertical="center")

# ===============================
#  ORDEN DE DOCENTES
# ===============================
def categoria_priority(categoria_id):
    if categoria_id not in [7, 8]:
        return 1
    elif categoria_id == 7:
        return 2
    elif categoria_id == 8:
        return 3

docentes_sorted = sorted(docentes_data, key=lambda d: categoria_priority(d["categoria_id"]))
otros = [d for d in docentes_sorted if categoria_priority(d["categoria_id"]) == 1]
tc    = [d for d in docentes_sorted if categoria_priority(d["categoria_id"]) == 2]
tp    = [d for d in docentes_sorted if categoria_priority(d["categoria_id"]) == 3]

# ===============================
#  FUNCIONES
# ===============================
def escribir_docentes(ws, docentes, row_start, contador, df, cargas_data):
    last_col = 2 + len(df) + 1
    fila_actual = row_start
    col_letter = ws.cell(row=2, column=last_col).column_letter  # Columna TOTAL

    for idx, docente in enumerate(docentes):
        fila1, fila2 = row_start + (idx * 2), row_start + (idx * 2) + 1
        nombre_completo = f"{docente['nombres']} {docente['apellidos']}"
        color_fondo = "FFFFFF" if contador % 2 != 0 else "DDDDDD"
        fill_style = PatternFill(start_color=color_fondo, end_color=color_fondo, fill_type="solid")

        # Columna A y B
        ws.merge_cells(f"A{fila1}:A{fila2}")
        ws[f"A{fila1}"] = contador
        ws[f"A{fila1}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"A{fila1}"].font = Font(bold=True, color="000000")

        ws.merge_cells(f"B{fila1}:B{fila2}")
        ws[f"B{fila1}"] = nombre_completo
        ws[f"B{fila1}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"B{fila1}"].font = Font(bold=True, color="000000")

        # Colorear filas completas
        for col in range(1, last_col + 1):
            ws.cell(row=fila1, column=col).fill = fill_style
            ws.cell(row=fila2, column=col).fill = fill_style

        # Cruce docente - curso - carga
        total_horas = 0
        for col_idx, curso in enumerate(df.itertuples(index=False), start=3):
            curso_id = curso.id
            carga = next((c for c in cargas_data if str(c["curso_id"]) == str(curso_id) and c["docente_id"] == docente["id"]), None)
            if carga:
                if int(carga["horas_t_carga"]) > 0:
                    cell_t = ws.cell(row=fila1, column=col_idx, value=f"{carga['horas_t_carga']}T")
                    cell_t.alignment = Alignment(horizontal="center", vertical="center")
                    cell_t.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                    cell_t.font = Font(bold=True)
                    total_horas += int(carga["horas_t_carga"])

                if int(carga["horas_p_carga"]) > 0:
                    cell_p = ws.cell(row=fila2, column=col_idx, value=f"{carga['horas_p_carga']}P")
                    cell_p.alignment = Alignment(horizontal="center", vertical="center")
                    cell_p.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                    cell_p.font = Font(bold=True)
                    total_horas += int(carga["horas_p_carga"])

        # 🔹 Combinar la columna TOTAL para este docente
        ws.merge_cells(f"{col_letter}{fila1}:{col_letter}{fila2}")
        total_cell = ws[f"{col_letter}{fila1}"]
        total_cell.value = total_horas
        total_cell.alignment = Alignment(horizontal="center", vertical="center")
        total_cell.font = Font(bold=True)

        contador += 1
        fila_actual = fila2 + 1

    return fila_actual, contador

def fila_separadora(ws, row, titulo, last_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    ws.cell(row=row, column=1, value=titulo).font = Font(bold=True, color="000000")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=1).fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

# ===============================
#  ESCRITURA FINAL
# ===============================
row_start, contador = 7, 1
row, contador = escribir_docentes(ws, otros, row_start, contador, df, cargas_data)
fila_separadora(ws, row, "JEFES DE PRÁCTICA TIEMPO COMPLETO", 2 + len(df) + 1); row += 1
row, contador = escribir_docentes(ws, tc, row, contador, df, cargas_data)
fila_separadora(ws, row, "JEFES DE PRÁCTICA TIEMPO PARCIAL", 2 + len(df) + 1); row += 1
row, contador = escribir_docentes(ws, tp, row, contador, df, cargas_data)

# Columna TOTAL
last_col = 2 + len(df) + 1
cell = ws.cell(row=2, column=last_col, value="TOTAL DE HORAS II SEMESTRE")
cell.alignment = Alignment(textRotation=90, horizontal="center", vertical="center")
cell.font = Font(bold=True)  # 🔹 Negrilla

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

max_row = ws.max_row
max_col = ws.max_column

for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
    for cell in row:
        cell.border = thin_border

# Datos de cursos
for idx, row in df.iterrows():
    col = 3 + idx
    ws.cell(row=3, column=col, value=f"T{row['horas_t']} P{row['horas_p']}").alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=4, column=col, value=row['n_estudiantes']).alignment = Alignment(horizontal="center", vertical="center")
    
    # === Fila de Grupos (amarillo + negrita) ===
    grupos_cell = ws.cell(row=5, column=col, value=row['n_grupos'])
    grupos_cell.alignment = Alignment(horizontal="center", vertical="center")
    grupos_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Amarillo
    grupos_cell.font = Font(bold=True, color="000000")  # Negrita negro

    # === Fila de Estudiantes por grupo (rosado claro) ===
    est_x_grupo = math.ceil(int(row['n_estudiantes']) / int(row['n_grupos'])) if int(row['n_grupos']) > 0 else 0
    est_cell = ws.cell(row=6, column=col, value=est_x_grupo)
    est_cell.alignment = Alignment(horizontal="center", vertical="center")
    est_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Rosado claro


# Ruta a la carpeta Descargas del usuario actual
downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")

output_file = os.path.join(downloads_folder, "cursos_semestre_2025_con_docentes.xlsx")
wb.save(output_file)
print(f"Archivo guardado en: {output_file}")
